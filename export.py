import json
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import OUTPUT_DIR


def _style_header(ws, columns):
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    for col_idx, (header, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"
    ws.freeze_panes = "A2"


def _write_rows(ws, rows, field_keys, link_columns=None):
    data_font = Font(name="Calibri", size=10)
    link_font = Font(name="Calibri", size=10, color="0563C1", underline="single")
    thin_border = Border(bottom=Side(style="thin", color="E0E0E0"))
    link_columns = link_columns or set()

    for row_idx, row_data in enumerate(rows, 2):
        bg = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid") if row_idx % 2 == 0 else None
        for col_idx, key in enumerate(field_keys, 1):
            value = row_data.get(key, "")
            if isinstance(value, list):
                value = ", ".join(str(v) for v in value)
            elif value is None:
                value = ""

            cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value else "")
            cell.font = data_font
            cell.border = thin_border
            if bg:
                cell.fill = bg
            if col_idx in link_columns and value and str(value).startswith("http"):
                cell.hyperlink = str(value)
                cell.font = link_font


def export_upwork_excel(jobs, filename=None):
    if not jobs:
        print("No Upwork jobs to export.")
        return None

    if not filename:
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        filename = f"Upwork_Jobs_{timestamp}.xlsx"

    # Parse skills JSON strings back to lists for display
    for job in jobs:
        if isinstance(job.get("skills"), str):
            try:
                job["skills"] = json.loads(job["skills"])
            except (json.JSONDecodeError, TypeError):
                job["skills"] = []

    columns = [
        ("Title", 40),
        ("Job URL", 35),
        ("Skills", 35),
        ("Budget Type", 12),
        ("Budget Min", 12),
        ("Budget Max", 12),
        ("Experience", 14),
        ("Client Rating", 12),
        ("Client Spent", 14),
        ("Client Hire Rate", 14),
        ("Client Country", 16),
        ("Proposals", 10),
        ("Posted", 16),
        ("Category", 25),
        ("Search Query", 20),
        ("Scraped At", 20),
    ]

    field_keys = [
        "title", "job_url", "skills", "budget_type", "budget_min", "budget_max",
        "experience_level", "client_rating", "client_total_spent", "client_hire_rate",
        "client_country", "num_proposals", "posted_date", "category",
        "search_query", "scraped_at",
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Upwork Jobs"

    _style_header(ws, columns)
    _write_rows(ws, jobs, field_keys, link_columns={2})

    # Summary sheet
    ws_stats = wb.create_sheet("Summary", 0)
    stats_fill = PatternFill(start_color="1B1B1B", end_color="1B1B1B", fill_type="solid")
    stats_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    val_font = Font(name="Calibri", size=14, bold=True, color="2563EB")

    labels = ["Total Jobs", "Fixed Price", "Hourly", "Avg Budget Min", "Top Category"]
    total = len(jobs)
    fixed = sum(1 for j in jobs if "fixed" in str(j.get("budget_type", "")).lower())
    hourly = sum(1 for j in jobs if "hourly" in str(j.get("budget_type", "")).lower())
    budgets = [j["budget_min"] for j in jobs if j.get("budget_min")]
    avg_budget = f"${sum(budgets) / len(budgets):,.0f}" if budgets else "N/A"

    cats = {}
    for j in jobs:
        c = j.get("category", "")
        if c:
            cats[c] = cats.get(c, 0) + 1
    top_cat = max(cats, key=cats.get) if cats else "N/A"

    values = [total, fixed, hourly, avg_budget, top_cat]

    for col, (label, val) in enumerate(zip(labels, values), 1):
        hcell = ws_stats.cell(row=1, column=col, value=label)
        hcell.fill = stats_fill
        hcell.font = stats_font
        hcell.alignment = Alignment(horizontal="center")
        vcell = ws_stats.cell(row=2, column=col, value=str(val))
        vcell.font = val_font
        vcell.alignment = Alignment(horizontal="center")
        ws_stats.column_dimensions[get_column_letter(col)].width = 22

    filepath = OUTPUT_DIR / filename
    wb.save(filepath)
    print(f"\nExported {total} Upwork jobs to {filepath}")
    return filepath


def export_fiverr_excel(gigs, filename=None):
    if not gigs:
        print("No Fiverr gigs to export.")
        return None

    if not filename:
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        filename = f"Fiverr_Gigs_{timestamp}.xlsx"

    for gig in gigs:
        if isinstance(gig.get("tags"), str):
            try:
                gig["tags"] = json.loads(gig["tags"])
            except (json.JSONDecodeError, TypeError):
                gig["tags"] = []

    columns = [
        ("Title", 45),
        ("Gig URL", 35),
        ("Seller", 18),
        ("Level", 12),
        ("Price From", 12),
        ("Rating", 10),
        ("Reviews", 10),
        ("Category", 25),
        ("Tags", 35),
        ("Search Query", 20),
        ("Scraped At", 20),
    ]

    field_keys = [
        "title", "gig_url", "seller_name", "seller_level",
        "price_min", "rating", "num_reviews", "category",
        "tags", "search_query", "scraped_at",
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fiverr Gigs"

    _style_header(ws, columns)
    _write_rows(ws, gigs, field_keys, link_columns={2})

    # Summary sheet
    ws_stats = wb.create_sheet("Summary", 0)
    stats_fill = PatternFill(start_color="1B1B1B", end_color="1B1B1B", fill_type="solid")
    stats_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    val_font = Font(name="Calibri", size=14, bold=True, color="1DBF73")

    labels = ["Total Gigs", "Avg Price", "Avg Rating", "Top Rated", "With Reviews"]
    total = len(gigs)
    prices = [g["price_min"] for g in gigs if g.get("price_min")]
    avg_price = f"${sum(prices) / len(prices):,.0f}" if prices else "N/A"
    ratings = [g["rating"] for g in gigs if g.get("rating")]
    avg_rating = f"{sum(ratings) / len(ratings):.1f}" if ratings else "N/A"
    top_rated = sum(1 for g in gigs if "top" in str(g.get("seller_level", "")).lower())
    with_reviews = sum(1 for g in gigs if g.get("num_reviews"))

    values = [total, avg_price, avg_rating, top_rated, with_reviews]

    for col, (label, val) in enumerate(zip(labels, values), 1):
        hcell = ws_stats.cell(row=1, column=col, value=label)
        hcell.fill = stats_fill
        hcell.font = stats_font
        hcell.alignment = Alignment(horizontal="center")
        vcell = ws_stats.cell(row=2, column=col, value=str(val))
        vcell.font = val_font
        vcell.alignment = Alignment(horizontal="center")
        ws_stats.column_dimensions[get_column_letter(col)].width = 22

    filepath = OUTPUT_DIR / filename
    wb.save(filepath)
    print(f"\nExported {total} Fiverr gigs to {filepath}")
    return filepath
